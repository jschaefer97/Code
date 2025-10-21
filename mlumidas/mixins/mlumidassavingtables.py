import os
import re
from typing import Any, Dict, List, Union, Optional

import pandas as pd


class MLUMidasSavingTablesMixin:
    """
    Save nested dictionaries of pandas DataFrames that summarize selected variables.

    Expected leaves are DataFrames produced by:
      - results_dict["varselect"]["selected_vars_quarters"][dict_name][crit][period]
        (rows=variables, cols=quarters)
      - results_dict["varselect"]["selected_vars_quarters_periods"][dict_name][crit]
        (rows=variables, cols=MultiIndex(period, quarter))

    The directory structure mirrors the nested dictionary structure.
    Each dict key becomes a folder; each DataFrame leaf is saved as:
      - Excel:  <prefix>_<title>.xlsx
      - HTML:   <prefix>_<title>.html
    where <prefix> is 'q_selected_table' or 'qp_selected_table' depending on the table type.
    """

    # ---------- Public API ----------

    def save_selected_tables(
        self,
        tables_dict: Dict[str, Any],
        file_path: str,
        title: str,
    ) -> List[str]:
        """
        Save all DataFrame leaves found in a nested tables dictionary to disk.

        Args:
            tables_dict: Nested dictionary whose leaves are pandas DataFrames.
            file_path:   Root directory where outputs should be saved (created if missing).
            title:       A short descriptor appended to filenames, e.g., 'gdp_nowcast_2020_2024'.

        Returns:
            A list of absolute file paths to the saved files (both .xlsx and .html).
        """
        if not isinstance(tables_dict, dict):
            raise TypeError("tables_dict must be a dictionary")

        saved: List[str] = []
        root = os.path.abspath(file_path)
        self._ensure_dir(root)

        self._save_tables_nested(
            node=tables_dict,
            base_dir=root,
            saved_paths=saved,
            title=title,
        )
        return saved

    # ---------- Internal helpers ----------

    def _save_tables_nested(
        self,
        node: Any,
        base_dir: str,
        saved_paths: List[str],
        title: str,
    ) -> None:
        """
        Recursively walk a nested structure of dicts/lists and save DataFrame leaves.
        Each dict key becomes a subfolder under base_dir.
        """
        if isinstance(node, dict):
            for key, child in node.items():
                folder = os.path.join(base_dir, self._sanitize_key(key))
                self._save_tables_nested(child, folder, saved_paths, title)
            return

        if isinstance(node, (list, tuple)):
            for idx, child in enumerate(node):
                folder = os.path.join(base_dir, f"{idx:03d}")
                self._save_tables_nested(child, folder, saved_paths, title)
            return

        # Leaf handlers
        if isinstance(node, pd.DataFrame):
            self._ensure_dir(base_dir)

            # Detect table type for filename prefix
            prefix = self._detect_prefix(node)

            # Build filenames
            safe_title = self._sanitize_key(title)
            base_name = f"{prefix}_{safe_title}" if safe_title else prefix
            xlsx_path = os.path.join(base_dir, f"{base_name}.xlsx")
            html_path = os.path.join(base_dir, f"{base_name}.html")

            # Save Excel (with merged title header)
            self._write_excel_with_header(node, xlsx_path, header=self._derive_header(prefix, title))
            saved_paths.append(xlsx_path)

            # Save HTML (with H3 title)
            self._write_html_with_header(node, html_path, header=self._derive_header(prefix, title))
            saved_paths.append(html_path)
            return

        # Unknown leaf type — create folder and drop a README for traceability
        self._ensure_dir(base_dir)
        info_path = os.path.join(base_dir, "README.txt")
        if not os.path.exists(info_path):
            with open(info_path, "w", encoding="utf-8") as f:
                f.write(
                    "This folder corresponds to a leaf in the tables dictionary that "
                    "was not a pandas DataFrame (or list/tuple of DataFrames).\n"
                )

    @staticmethod
    def _derive_header(prefix: str, title: str) -> str:
        """Human-friendly header line written into Excel/HTML outputs."""
        label = "Quarter × Period Selected Table" if prefix.startswith("qp_") else "Quarter Selected Table"
        return f"{label} — {title}" if title else label

    @staticmethod
    def _detect_prefix(df: pd.DataFrame) -> str:
        """
        Decide filename prefix:
          - 'qp_selected_table' if columns are MultiIndex with names ['period','quarter']
          - otherwise 'q_selected_table'
        """
        if isinstance(df.columns, pd.MultiIndex):
            names = list(df.columns.names) if df.columns.names is not None else []
            names_lc = [str(n).lower() if n is not None else "" for n in names]
            if len(names_lc) >= 2 and "period" in names_lc[0] and "quarter" in names_lc[1]:
                return "qp_st"
        return "q_st"

    @staticmethod
    def _write_excel_with_header(df: pd.DataFrame, path: str, header: str) -> None:
        """
        Write DataFrame to Excel with a merged, bold header row above the table.
        Uses openpyxl under the hood (default engine for .xlsx).
        """
        # Write DataFrame starting at row 2 (1-based), so we can add header at row 1.
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Sheet1", startrow=1)  # leave row 1 for header
            ws = writer.sheets["Sheet1"]

            # Number of visible columns including index col
            n_cols = df.shape[1] + 1  # +1 for index column
            try:
                from openpyxl.styles import Font, Alignment
            except Exception:
                # If styles aren't available, still try to write the header text
                n_cols = max(n_cols, 1)
                ws.cell(row=1, column=1, value=header)
                return

            # Merge A1 : <last_col>1
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            cell = ws.cell(row=1, column=1, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _write_html_with_header(df: pd.DataFrame, path: str, header: str) -> None:
        """
        Write DataFrame to a simple HTML file with an H3 header and minimal styling.
        """
        table_html = df.to_html(border=0, classes="table", notebook=False)
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{header}</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 20px; }}
  h3 {{ margin-bottom: 12px; }}
  table {{ border-collapse: collapse; font-size: 13px; }}
  th, td {{ border: 1px solid #ddd; padding: 6px 8px; }}
  th {{ background: #f5f5f5; }}
  tfoot td {{ font-weight: bold; }}
</style>
</head>
<body>
  <h3>{header}</h3>
  {table_html}
</body>
</html>"""
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)

    @staticmethod
    def _ensure_dir(path: str) -> None:
        """Create a directory if it doesn't exist."""
        os.makedirs(path, exist_ok=True)

    @staticmethod
    def _sanitize_key(key: Union[str, int, float, None]) -> str:
        """
        Convert arbitrary dictionary keys (e.g., periods, timestamps) to safe folder names.
        """
        if key is None:
            s = "None"
        else:
            s = str(key)

        # Replace path separators and trim whitespace
        s = s.strip().replace(os.sep, "_").replace("/", "_").replace("\\", "_")

        # Collapse spaces and disallowed characters
        s = re.sub(r"\s+", "_", s)
        s = re.sub(r"[^A-Za-z0-9._\-+=@]", "_", s)

        # Avoid empty names
        return s or "key"
